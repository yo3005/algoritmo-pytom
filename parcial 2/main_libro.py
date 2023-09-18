# from openpyxl import Workbook, load_workbook
import pandas as pd
import openpyxl
from tabulate import tabulate
from libro_negocio import LibroNegocio

# crud
# crear
# leer
# actualizar
# eliminar
##########################
listado_libro=[]
#region autores
negocio = LibroNegocio()

# Funcion que muestra todos los autores registrados
def listar_autores():
    print('\n\t\t\t\t\t     autores registrados\n'.center(250,'='))
    # codigo para guargar doatos del archivo listado autores y mostrarla en una tabla con tabulate
    # ----------------------------------------------------------------
    excel_dataframe = openpyxl.load_workbook('listado_autores.xlsx')
    dataframe=excel_dataframe.active
    
    data = []
    
    for row in range(1,dataframe.max_row):
        
        _row = [row,]
        for column in dataframe.iter_cols(1,dataframe.max_column):
            _row.append(column[row].value)
        data.append(_row)
    headers = ['indice','Nombre', 'Apellido Paterno', 'Apellido Materno', 'Fecha de Nacimiento', 'Codigo del Autor', 'Pais', 'Editorial']
    print(tabulate(data,headers=headers))
    # ----------------------------------------------------------------
    print(''.center(109,'='))


def listar_categorias():
    print('\n\t\t\t\t\t     Categorias registradas\n'.center(250,'='))
    # codigo para guargar doatos del archivo listado autores y mostrarla en una tabla con tabulate
    # ----------------------------------------------------------------
    excel_dataframe = openpyxl.load_workbook('listado_categorias.xlsx')
    dataframe=excel_dataframe.active
    
    data = []
    
    for row in range(1,dataframe.max_row):
        
        _row = [row,]
        for column in dataframe.iter_cols(1,dataframe.max_column):
            _row.append(column[row].value)
        data.append(_row)
    headers = ['indice','Codigo de la Categoria', 'Categoria']
    print(tabulate(data,headers=headers))
    # ----------------------------------------------------------------
    print(''.center(109,'='))


def registrar_libros():
    # codigo_libro, titulo, anio, codigo_autor, codigo_categoria
    codigo_libro = input('\nIngrese codigo para el libro: ')
    titulo = input('Ingrese titulo del libro: ')
    anio = input('Ingrese anio del libro: ')
    
    listar_autores()
    codigo_autor = input('\nIngrese el codigo del autor a asignar al libro: ')
    
    listar_categorias()
    codigo_categoria = input('\nIngrese el codigo de la categoria a asignar al libro: ')

    
    negocio.registrar_libros(codigo_libro, titulo, anio, codigo_autor, codigo_categoria)
    negocio.guardar_libros()  


# Fucin que meustra todos los libros registrados
def obtener_libros():
    
    print('\n')
    print('\n \t\t\t\tlibros registrados\n'.center(189,'='))
    # codigo para guargar doatos del archivo listado libros y mostrarla en una tabla con tabulate
    # ----------------------------------------------------------------
    excel_dataframe = openpyxl.load_workbook('listado_libros.xlsx')
    dataframe=excel_dataframe.active
    
    data = []
    
    for row in range(1,dataframe.max_row):
        
        _row = [row,]
        for column in dataframe.iter_cols(1,dataframe.max_column):
            _row.append(column[row].value)
        data.append(_row)
    headers = ['indice','Codigo del Libro', 'Titulo', 'Año', 'Codigo autor', 'Nombre del Autor', 'Codigo de la categoria', 'Categoria']
    print(tabulate(data,headers=headers),'\n')
    # ----------------------------------------------------------------
    print(''.center(82,'='))
    

# Funcion que permite editar libros
def editar_libro():
    
    # Muestra los libros registrados
    obtener_libros()
    
    indice = int(input('\nIngrese el indice del libro a editar: ')) - 1
    codigo_libro = input('\nIngrese el codigo del nuevo curos: ')
    titulo = input('Ingrese el titulo del nuevo libro: ')
    anio = input('Ingrese el año del nuevo libro: ')
    
    # Muestra los autores registrados
    listar_autores()
    codigo_autor = input('Ingrese el codigo del nuevo autor: ')
    
    listar_categorias()
    codigo_categoria = input('Ingrese el codigo de la nueva categoria: ')
    
    print(negocio.editar_libros(indice, codigo_libro, titulo, anio, codigo_autor, codigo_categoria))


def eliminar_libro():
    
    #  Muestra los libros registrados
    obtener_libros()
    
    indice_a_eliminar = int(input("\nIngrese el indice del autor a eliminar: ")) - 1
    print(negocio.eliminar_libros(indice_a_eliminar))
#endregion


##diccionario
opciones = {
    "1": registrar_libros,
    "2": obtener_libros,
    "3": editar_libro,
    "4": eliminar_libro,
    "5": exit
}

while True:
    print('\n')
    print("\t\n\t\tMenú\n".center(80,'='))
    print("1. Rgistrar libros")
    print("2. Mostrar libros")
    print("3. Editar libro")
    print("4. Eliminar libro")
    print("5. Salir")
    print("".center(36,'='))
    
    seleccion = input("Seleccione una opción: ")

    if seleccion in opciones:
        opciones[seleccion]()
    else:
        print("Opción no válida. Por favor, seleccione una opción válida.")