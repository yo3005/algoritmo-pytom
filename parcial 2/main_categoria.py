import pandas as pd
import openpyxl
from openpyxl import Workbook
from tabulate import tabulate
from categoria_negocio import CategoriaNegocio

listado_categoria=[]

#region categorias
negocio = CategoriaNegocio()


def registrar_categorias():
    codigo_categoria = input('\nIngrese Codigo Categoria: ')
    categoria = input('Ingrese nombre de la Categoria: ')
    negocio.registrar_categorias(codigo_categoria,categoria)
    negocio.guardar_categorias()


def obtener_categorias():
    
    print('\n')
    print('\n \t\t\t\t\t\tcategorias registradas\n'.center(247,'='))
    
    # Mostrando todos los categorias registrados
    # ----------------------------------------------------------------
    excel_dataframe = openpyxl.load_workbook('listado_categorias.xlsx')
    dataframe=excel_dataframe.active
    
    data = []
    
    for row in range(1,dataframe.max_row):
        
        _row = [row,]
        for column in dataframe.iter_cols(1,dataframe.max_column):
            _row.append(column[row].value)
        data.append(_row)
    headers = ['indice','Codigo de la Categoria', 'Categoria']
    # tablefmt='fancy_grid'
    print(tabulate(data,headers=headers))
    # ----------------------------------------------------------------
    print(''.center(110,'='))

# Funcion que permite editar categoria
def editar_categoria():
    
    obtener_categorias()
    indice = int(input('\nIngrese el indice del categoria a editar: ')) - 1
    codigo_categoria = int(input('\nIngrese Codigo Categoria: '))
    categoria = input('Ingrese nombre de la Categoria: ')

    print(negocio.editar_categorias(indice,codigo_categoria,categoria))


# funcion que permite eliminar categoria
def eliminar_categoria():
    
    obtener_categorias()
    
    indice_a_eliminar = int(input("\nIngrese el indice del categoria a eliminar: ")) - 1
    print(negocio.eliminar_categoria(indice_a_eliminar))
#endregion

##diccionario
opciones = {
    "1": registrar_categorias,
    "2": obtener_categorias,
    "3": editar_categoria,
    "4": eliminar_categoria,
    "5": exit
}

while True:
    print("\n\t\tMenú\n".center(80,'='))
    print("1. registrar categorias")
    print("2. obtener categorias")
    print("3. editar categoria")
    print("4. eliminar categoria")
    print("5. Salir")
    print("".center(36,'='))
    
    seleccion = input("Seleccione una opción: ")

    if seleccion in opciones:
        opciones[seleccion]()
    else:
        print("Opción no válida. Por favor, seleccione una opción válida.")