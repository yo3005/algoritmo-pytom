import pandas as pd
import openpyxl
from tabulate import tabulate
from autor_negocio import  AutorNegocio
from datetime import datetime

listado_autores=[]

#region autores
negocio = AutorNegocio()


def registrar_autores():
    nombre = input('Ingrese Nombre: ')
    apellido_paterno = input('Ingrese Apellido Paterno: ')
    apellido_materno = input('Ingrese Apellido Materno: ')
    fecha_nacimiento = input('Ingrese Fecha de Nacimiento: ')
    codigo_autor = input('Ingrese Codigo del Autor: ')
    pais = input('Ingrese Pais: ')
    editorial = input('Ingrese Editorial: ')
    negocio.registrar_autores(nombre, apellido_paterno, apellido_materno, fecha_nacimiento, codigo_autor, pais, editorial)
    negocio.guardar_autores()
    print(f'registro correctamente elestudiantes')

def obtener_autores():
    print('\n')
    print('\n \t\t\t\t\t\t\tCursos registrados\n'.center(276,'='))
    
    # ----------------------------------------------------------------
    excel_dataframe = openpyxl.load_workbook('listado_autores.xlsx')
    dataframe=excel_dataframe.active
    
    data = []
    
    for row in range(1,dataframe.max_row):
        
        _row = [row,]
        for column in dataframe.iter_cols(1,dataframe.max_column):
            _row.append(column[row].value)
        data.append(_row)
    headers = ['indice','Nombre', 'Apellido Paterno', 'Apellido Materno', 'Fecha de Nacimiento', 'Codigo del Autor', 'Pais', 'Editorial']
    
    # print(tabulate(data,headers=headers),'\n')
    print(tabulate(data,headers=headers),'\n')
    # ----------------------------------------------------------------
    print(''.center(124,'='))

def editar_autor():
    
    obtener_autores()
    indice = int(input('Ingrese el indice a editar: '))
    nombre = input('Ingrese Nombre: ')
    apellido_paterno = input('Ingrese Apellido Paterno: ')
    apellido_materno = input('Ingrese Apellido Materno: ')
    fecha_nacimiento = input('Ingrese Fecha de Nacimiento: ')
    codigo_autor = input('Ingrese Codigo del Autor: ')
    pais = input('Ingrese Pais: ')
    editorial = input('Ingrese Editorial: ')
    print(negocio.editar_autor(indice - 1, nombre, apellido_paterno, apellido_materno, fecha_nacimiento, codigo_autor, pais,editorial))

def eliminar_autor():
    
    #  Muestra los cursos registrados
    obtener_autores()
    
    indice_a_eliminar = int(input("\nIngrese el indice del autor a eliminar: ")) - 1
    print(negocio.eliminar_autor(indice_a_eliminar))
#endregion


##diccionario
opciones = {
    "1": registrar_autores,
    "2": obtener_autores,
    "3": editar_autor,
    "4": eliminar_autor,
    "5": exit
}

while True:
    print("\n\t\tMenú\n".center(80,'='))
    print("1. Registrar autores")
    print("2. Listar autores")
    print("3. Editar autor")
    print("4. Eliminar autor")
    print("5. Salir")
    print("".center(36,'='))    
    seleccion = input("Seleccione una opción: ")

    if seleccion in opciones:
        opciones[seleccion]()
    else:
        print("Opción no válida. Por favor, seleccione una opción válida.")